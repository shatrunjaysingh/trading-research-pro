"""
Daily report — fetches 3 research views, builds a styled Excel workbook,
and emails it to the configured recipient.

Sheets:
  1. Top 20 Stocks       — all prices
  2. Stocks Under $5     — price-filtered stocks
  3. Top 20 Crypto       — all crypto

Usage:
  python3 report.py

Required env vars (set once in ~/.zshrc):
  EMAIL_SENDER        your Gmail address
  EMAIL_APP_PASSWORD  Gmail App Password (not your login password)
                      → myaccount.google.com → Security → App passwords
"""

import os
import io
import logging
import smtplib
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pytz
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from research import fetch_all_data, fetch_cheap_stocks, load_config, run_api, PICKS_PER_CATEGORY

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s — %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
BUY_FILL      = PatternFill("solid", fgColor="C6EFCE")   # light green
WATCH_FILL    = PatternFill("solid", fgColor="DDEBF7")   # light blue
HOLD_FILL     = PatternFill("solid", fgColor="FFEB9C")   # light yellow
ALT_FILL      = PatternFill("solid", fgColor="F2F2F2")   # light grey for alternating rows
THIN_BORDER   = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

COLUMNS = [
    ("Rank",           6),
    ("Ticker",         9),
    ("Type",           8),
    ("Score",          7),
    ("Conf %",         7),
    ("Signal",         8),
    ("Price (USD)",   13),
    ("Day %",          8),
    ("3M %",           8),
    ("Vol Ratio",      9),
    ("52w Pos",        9),
    ("Analyst Target",13),
    ("Upside %",       9),
    ("Consensus",     13),
    ("Insider",       14),
    ("Inst Own %",    10),
    ("Earnings",      10),
]


def _insider_label(r: dict) -> str:
    net = r.get("insider_net_shares", 0)
    if net > 20_000:
        return f"Buying (+{net:,})"
    if net < -50_000:
        return f"Selling ({net:,})"
    return "Neutral"


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def _write_sheet(ws, rows: list, title: str, subtitle: str) -> None:
    """Write one sheet: title row, header row, data rows with styling."""

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Subtitle / date row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font      = Font(italic=True, size=10, color="595959")
    sub_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Header row
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[3].height = 18

    # Data rows
    signal_fill = {"BUY": BUY_FILL, "WATCH": WATCH_FILL, "HOLD": HOLD_FILL}

    for row_idx, r in enumerate(rows, start=1):
        excel_row = row_idx + 3
        fill = signal_fill.get(r["signal"], ALT_FILL) if row_idx % 2 == 1 else ALT_FILL

        target  = r.get("analyst_target")
        upside  = r.get("analyst_upside_pct", 0)
        values = [
            row_idx,
            r["ticker"],
            r.get("type", "").capitalize(),
            r["score"],
            f"{r.get('confidence', 0)}%",
            r["signal"],
            r["current_price"],
            f"{r['day_change_pct']:+.2f}%",
            f"{r.get('qtr_change_pct', 0):+.2f}%",
            f"{r['vol_ratio']:.2f}x",
            f"{r['pos_52w']:.1f}%",
            f"${target:.2f}" if target else "—",
            f"{upside:+.1f}%" if target else "—",
            r.get("analyst_consensus", "—") if r.get("num_analysts", 0) >= 2 else "—",
            _insider_label(r),
            f"{r.get('inst_pct_held', 0):.0f}%" if r.get("inst_pct_held") else "—",
            r.get("earnings_flag") or "—",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell            = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill       = fill
            cell.border     = THIN_BORDER
            cell.alignment  = Alignment(horizontal="center")

        ws.row_dimensions[excel_row].height = 16

    # Freeze panes below header
    ws.freeze_panes = ws["A4"]

    # Auto-filter on header row
    ws.auto_filter.ref = (
        f"A3:{get_column_letter(len(COLUMNS))}3"
    )


def build_workbook(stock_rows, cheap_rows, crypto_rows, run_date: str) -> Workbook:
    wb = Workbook()

    sheets = [
        (wb.active,       "Top 20 Stocks",    stock_rows,  "All stocks, ranked by momentum + volume + 52-week position"),
        (wb.create_sheet("Stocks Under $5"),  "Stocks Under $5",  cheap_rows,  "Stocks priced below $5, ranked by score"),
        (wb.create_sheet("Top 20 Crypto"),    "Top 20 Crypto",    crypto_rows, "Crypto assets, ranked by momentum + volume + 52-week position"),
    ]

    # First sheet is already named; rename it
    wb.active.title = "Top 20 Stocks"

    for ws, title, rows, subtitle_text in [
        (wb["Top 20 Stocks"],  "Top 20 Stocks",   stock_rows,  f"Stocks — {run_date}  |  {subtitle_text}" ),
        (wb["Stocks Under $5"],"Stocks Under $5", cheap_rows,  f"Price ≤ $5 — {run_date}  |  {subtitle_text}"),
        (wb["Top 20 Crypto"],  "Top 20 Crypto",   crypto_rows, f"Crypto — {run_date}  |  {subtitle_text}"),
    ]:
        _write_sheet(ws, rows, title, f"{run_date}  |  {subtitle_text}")

    # Fix: sheets were defined twice — simplify
    return wb


def build_workbook(stock_rows: list, cheap_rows: list, crypto_rows: list, run_date: str) -> Workbook:
    wb = Workbook()
    wb.active.title = "Top 20 Stocks"
    wb.create_sheet("Stocks Under $5")
    wb.create_sheet("Top 20 Crypto")

    score_note = "Score = momentum(40%) + volume surge(30%) + 52-week position(30%)"

    _write_sheet(wb["Top 20 Stocks"],
                 stock_rows,
                 "Top 20 Stocks",
                 f"{run_date}  |  {score_note}")

    _write_sheet(wb["Stocks Under $5"],
                 cheap_rows,
                 "Stocks Under $5",
                 f"Price ≤ $5.00  |  {run_date}  |  {score_note}")

    _write_sheet(wb["Top 20 Crypto"],
                 crypto_rows,
                 "Top 20 Crypto",
                 f"{run_date}  |  {score_note}")

    return wb


# ---------------------------------------------------------------------------
# Email sender
# ---------------------------------------------------------------------------

def send_email(wb: Workbook, config: dict, run_date: str) -> None:
    email_cfg   = config["email"]
    raw         = email_cfg["recipient"]
    recipients  = raw if isinstance(raw, list) else [raw]
    sender      = os.getenv("EMAIL_SENDER") or email_cfg.get("sender", "")
    # Strip spaces/non-breaking spaces Google inserts when displaying App Passwords
    app_password = "".join(os.getenv("EMAIL_APP_PASSWORD", "").split())
    smtp_host   = email_cfg.get("smtp_host", "smtp.gmail.com")
    smtp_port   = email_cfg.get("smtp_port", 587)

    if not sender:
        logger.error("EMAIL_SENDER env var not set. Export it and re-run.")
        raise SystemExit(1)
    if not app_password:
        logger.error("EMAIL_APP_PASSWORD env var not set. Export it and re-run.")
        raise SystemExit(1)

    # Serialize workbook to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"trading_picks_{run_date.replace(', ', '_').replace(' ', '_')}.xlsx"

    # Build email
    msg = MIMEMultipart()
    msg["From"]    = sender
    msg["To"]      = ", ".join(recipients)
    msg["Subject"] = f"Daily Trading Picks - {run_date}"

    body = (
        f"Hi,\n\n"
        f"Please find today's trading research attached ({run_date}).\n\n"
        f"The workbook contains three sheets:\n"
        f"  - Top 20 Stocks      : all prices\n"
        f"  - Stocks Under $5    : price-filtered stocks\n"
        f"  - Top 20 Crypto      : cryptocurrency picks\n\n"
        f"Scoring: momentum (40%) + volume surge (30%) + 52-week position (30%)\n"
        f"Signals: BUY (score >= 65) | WATCH (45-64) | HOLD (< 45)\n\n"
        f"NOTE: This is automated research only - no trades have been placed.\n\n"
        f"- Trading Agent"
    )
    msg.attach(MIMEText(body, "plain"))

    # Attach Excel
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(buf.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    # Send
    logger.info("Sending email to %s via %s:%d ...", ", ".join(recipients), smtp_host, smtp_port)
    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.login(sender, app_password)
        server.sendmail(sender, recipients, msg.as_string())

    logger.info("Email sent to %s — %s", ", ".join(recipients), filename)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    config      = load_config()
    et          = pytz.timezone("America/New_York")
    now         = datetime.now(et)
    run_date    = now.strftime("%A, %Y-%m-%d %H:%M ET")
    mode        = config.get("research", {}).get("mode", "free")
    top_n       = config.get("research", {}).get("top_n", PICKS_PER_CATEGORY)
    stocks      = config["assets"]["stocks"]
    crypto      = config["assets"]["crypto"]
    email_cfg   = config.get("email")

    # ── API (deep research) mode: Claude + web search, rich HTML email ──────────
    if mode == "api":
        if not os.getenv("ANTHROPIC_API_KEY"):
            logger.warning("mode=api but ANTHROPIC_API_KEY not set — falling back to free mode.")
        else:
            logger.info("Deep research mode (Claude API) — running dual-category analysis...")
            run_api(stocks, crypto, top_n=top_n, dual_category=True, email_cfg=email_cfg)
            return

    # ── Free mode: yfinance multi-factor scoring, Excel workbook ────────────────
    logger.info("Fetching market data for %d assets...", len(stocks) + len(crypto))
    all_rows = fetch_all_data(stocks, crypto)

    if not all_rows:
        logger.error("No data fetched — aborting.")
        return

    stock_rows = sorted(
        [r for r in all_rows if r["type"] == "stock"],
        key=lambda x: x["score"], reverse=True
    )[:20]

    cheap_rows = fetch_cheap_stocks(max_price=5.0, min_market_cap=10_000_000, limit=50)[:20]

    crypto_rows = sorted(
        [r for r in all_rows if r["type"] == "crypto"],
        key=lambda x: x["score"], reverse=True
    )[:20]

    logger.info("Stocks: %d picks | Under $5: %d picks | Crypto: %d picks",
                len(stock_rows), len(cheap_rows), len(crypto_rows))

    wb = build_workbook(stock_rows, cheap_rows, crypto_rows, run_date)
    logger.info("Workbook built with 3 sheets.")
    send_email(wb, config, run_date)


if __name__ == "__main__":
    main()
